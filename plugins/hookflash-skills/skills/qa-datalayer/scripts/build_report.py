#!/usr/bin/env python3
"""Build the qa-datalayer .xlsx report.

Usage:  python build_report.py events.json <screenshots_dir> <out.xlsx>

Keep <out.xlsx> SHORT and generic (e.g. CB_QA.xlsx). The session output directory is
already ~200 chars; a long filename breaks the Windows 259-char path limit and the
workbook will not open. Put the descriptive title inside the workbook, not in the name.

events.json = list of objects, one per event, in funnel order:
{
  "event": "add_to_cart",
  "source": "Top frame",                 # "Top frame" | "Web pixel" | "Not fired"
  "conditions": "Clicked 'Add to basket' on the PDP.",
  "push": { ... captured object, verbatim ... },   # dict -> pretty JSON; or null
  "push_note": "No event fired on click.",           # used ONLY when push is null
  "location_image": "atc_button.png",                # filename inside screenshots_dir
  "location_bbox": [x, y, w, h],                      # optional fallback: crop to element
  "viewport_w": 1280,                                 # image width the bbox is expressed in
  "verdict": "pass",                                  # pass | fail | warn | na
  "notes": ["- items array = single added item", "- core params populated"]
}

Verdict / notes style (keep the Pass/Fail column skimmable):
- One short bullet per REAL finding. Lead with an "All expected parameters present"
  line when coverage is complete, then list only the problems.
- Do NOT narrate non-issues (params legitimately null pre-cart, standard params absent
  but not in the spec, item-schema differences that are just null/absent fields).
- Bullets are rendered with a blank line between them automatically (below), so each
  finding stays on its own line instead of congealing into one block.

Rules baked in here:
- The "push" dict is dumped VERBATIM with json.dumps(indent=2). Do not pre-format it.
- For a Web pixel event whose runtime values cannot be read, build "push" as the object
  shape from the pixel source, each unreadable value set to "(Can't read values in web pixel)".
- Em/en dashes are stripped from all prose cells (never from the JSON).
- Images are auto-fit to the screenshot column (IMG_W < column E width) so they never
  overhang into the Pass/Fail column, and each row is tall enough for its image.
"""
import json, os, sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
try:
    from PIL import Image as PILImage
    HAVE_PIL = True
except Exception:
    HAVE_PIL = False

DASHES = {"—": "-", "–": "-", "‒": "-", "―": "-", "−": "-"}
def nodash(s):
    if not isinstance(s, str):
        s = str(s)
    for k, v in DASHES.items():
        s = s.replace(k, v)
    return s

def main():
    if len(sys.argv) < 4:
        print(__doc__); sys.exit(1)
    events_path, shots_dir, out_path = sys.argv[1], sys.argv[2], sys.argv[3]
    events = json.load(open(events_path, encoding="utf-8"))

    FONT, MONO = "Arial", "Consolas"
    hdr = PatternFill("solid", fgColor="1F2A44")
    fills = {"pass": PatternFill("solid", fgColor="D6EAD6"),
             "fail": PatternFill("solid", fgColor="F7D6D6"),
             "warn": PatternFill("solid", fgColor="FCEFC7"),
             "na":   PatternFill("solid", fgColor="E6E6E6")}
    src_fills = {"top frame": PatternFill("solid", fgColor="DCE7F5"),
                 "web pixel": PatternFill("solid", fgColor="EADCF5"),
                 "not fired": PatternFill("solid", fgColor="EEEEEE")}
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    top = Alignment(horizontal="left", vertical="top", wrap_text=True)

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "DataLayer QA"
    headers = ["Event name", "Source", "Conditions tested",
               "dataLayer push (verbatim JSON)", "Location screenshot", "Pass / Fail"]
    # A widened so event names (e.g. begin_checkout, view_item_list) fit on one line.
    # E kept wider than IMG_W (350px) so images never overhang into F.
    widths = {"A": 20, "B": 13, "C": 34, "D": 60, "E": 54, "F": 46}
    for k, v in widths.items():
        ws.column_dimensions[k].width = v
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
        cell.fill = hdr; cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    IMG_W = 350
    r = 2
    for ev in events:
        push = ev.get("push", None)
        if push is not None:
            js = json.dumps(push, indent=2, ensure_ascii=False)   # verbatim, NOT dash-stripped
        else:
            js = nodash(ev.get("push_note", "(no event captured)"))
        notes = ev.get("notes", [])
        # blank line between bullets so findings stay visually separated
        notes_txt = nodash("\n\n".join(notes) if isinstance(notes, list) else str(notes))
        source = ev.get("source", "")

        ws.cell(row=r, column=1, value=nodash(ev.get("event", ""))).font = Font(name=FONT, size=11, bold=True)
        ws.cell(row=r, column=2, value=nodash(source)).font = Font(name=FONT, size=10, bold=True)
        ws.cell(row=r, column=3, value=nodash(ev.get("conditions", ""))).font = Font(name=FONT, size=10)
        ws.cell(row=r, column=4, value=js).font = Font(name=MONO, size=9)
        ws.cell(row=r, column=6, value=notes_txt).font = Font(name=FONT, size=10)
        for c in range(1, 7):
            cc = ws.cell(row=r, column=c); cc.alignment = top; cc.border = border
        vfill = fills.get(ev.get("verdict", "warn"))
        ws.cell(row=r, column=1).fill = vfill
        ws.cell(row=r, column=6).fill = vfill
        sfill = src_fills.get(source.strip().lower())
        if sfill:
            ws.cell(row=r, column=2).fill = sfill

        img_h = 0
        img_file = ev.get("location_image")
        if img_file and str(img_file).strip().lower() not in ("none", "null", ""):
            path = os.path.join(shots_dir, img_file)
            if os.path.exists(path) and HAVE_PIL:
                if ev.get("location_bbox"):
                    x, y, w, h = ev["location_bbox"]
                    try:
                        with PILImage.open(path) as im:
                            sx = im.width / ev.get("viewport_w", im.width)
                            crop = im.crop((int(x*sx), int(y*sx), int((x+w)*sx), int((y+h)*sx)))
                            cpath = path + ".crop.png"; crop.save(cpath); path = cpath
                    except Exception:
                        pass
                with PILImage.open(path) as im:
                    w0, h0 = im.size
                img = XLImage(path); img.width = IMG_W; img.height = int(IMG_W * h0 / w0)
                ws.add_image(img, "E%d" % r); img_h = img.height

        js_pts = (js.count("\n") + 1) * 12
        notes_pts = (notes_txt.count("\n") + 1) * 14
        img_pts = img_h * 0.75           # px -> pt, so the row is tall enough for the image
        ws.row_dimensions[r].height = max(js_pts, notes_pts, img_pts + 14, 120)
        r += 1

    wb.save(out_path)
    if len(os.path.abspath(out_path)) >= 259:
        print("WARNING: output path is >= 259 chars, Windows Excel may refuse to open it. Use a shorter filename.")
    print("saved", out_path, "rows", r-2)

if __name__ == "__main__":
    main()
