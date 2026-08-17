#!/usr/bin/env python3
"""Build the qa-network-requests .xlsx report.

Usage:  python build_report.py events.json consent.json <screenshots_dir> <out.xlsx>

Workbook layout:
  Tab 1        "Consent"  - always first, descriptive only, never graded
  Tab 2..N     one tab PER PLATFORM (GA4, Meta, TikTok, ...), in the order the
               vendors first appear in events.json

consent.json may be the literal string "none", but prefer a real file: the consent
check is forced now (cookies cleared + reload), so there is always something to say.

Keep <out.xlsx> SHORT and generic (e.g. CB_NQA.xlsx). The session output directory is
already ~200 chars; a long filename breaks the Windows 259-char path limit and the
workbook will not open. Put the descriptive title inside the workbook, not in the name.

events.json = list of objects, one per HIT, in funnel order. `vendor` decides which
tab it lands on, so it is required:
{
  "event": "nav_click",                    # the name the SPEC uses
  "vendor": "GA4",                         # -> tab name
  "sent": true,                            # false -> "No request observed"
  "method": "POST",
  "endpoint": "metrics.example.com/g/collect",   # host + path only
  "url": "https://metrics.example.com/g/collect?v=2&tid=G-XXX&en=nav_click...",  # VERBATIM
  "body": "en=nav_click&ep.link_text=Shop",      # verbatim body, or null
  "payload": { ... every decoded param ... },    # -> "Full payload" column

  # --- the "Spec parameters" column: ONLY what the spec asked about ---
  "spec_params": ["en", "ep.link_text", "ep.link_url"],
        # Names as THIS VENDOR sends them (you just decoded them, so you know).
        # Values are pulled out of "payload" above, so the two columns can never
        # disagree, and anything the spec wants but the hit lacks renders as
        # "(absent)" - which is the single most useful cell in the report.
        # Supported forms:
        #   "en"                  literal key (dots are fine: "epn.value", "cd[value]")
        #   "items[].item_id"     item-level param, returns one value per item
        #   "__body_json.properties.value"   dotted path into a JSON body
  "spec_payload": { ... },        # optional explicit override, wins over spec_params.
                                  # Use when the spec-to-vendor mapping is not a lookup.

  "absence_evidence": "68 requests searched via __nqaRequests(), none matching /g/collect.",
        # REQUIRED whenever "sent" is false. Without it the cell is stamped
        # "!! UNVERIFIED" and the run prints a warning, because a filtered read
        # coming back empty is not evidence that nothing was sent.

  "count": 2,                              # optional; rendered when > 1 (duplicate tagging)
  "conditions": "Clicked 'Shop' nav link, header",
        # One short phrase. What was done, not why, not what it proves.
  "location_image": "nav_shop.png",        # filename inside screenshots_dir
  "verdict": "fail",                       # pass | fail | warn | na
  "notes": ["- Event name, tid match spec", "- ep.link_url absent"]
        # Terse bullets, observation only, max 5. The reader is an analyst who
        # knows GA4/Meta/TikTok: name the param and what is wrong with it, and
        # stop. No explanations of what a param is for or why it matters.
}

consent.json = list of objects, one per vendor, optionally preceded by a state object:
{"state": "forced", "reset": "Cleared 14 cookies + localStorage, then reloaded."}
{
  "vendor": "Meta",
  "before": "No hits observed",
  "evidence": "41 requests searched via __nqaRequests(), none matching facebook.com/tr.",
        # REQUIRED when "before" claims nothing fired, same gate as absence_evidence.
  "signal": "n/a (no consent parameter)",
  "after": "1 hit (PageView)",
  "observation": "Gated on consent. First /tr hit is PageView, post-accept.",
        # One or two clauses, descriptive, never a verdict.
  "location_image": "banner.png"
}

Design rules baked in here:
- Columns are WIDE and prose cells are SHORT, because row height is what makes this
  workbook unreadable. Cells you write ("conditions", "notes", "observation") are
  observations for an analyst, not explanations: overrunning the MAX_* budgets below
  is reported at the end of the run. Evidence cells have no budget and are never cut.
- "url" and "body" are rendered VERBATIM, character for character. Never pre-format,
  trim or tidy them. Payload VALUES are equally verbatim; only their layout is ours
  (params are packed two per line to keep the row short, see payload_text).
- "Spec parameters" is DERIVED from "payload" by this script, not retyped, so it
  cannot drift from the evidence beside it.
- The Consent tab has NO pass/fail column, deliberately, and comes first. Consent
  behaviour is reported, not graded. Do not add a verdict column to it.
- Em/en dashes are stripped from prose cells (never from URLs, bodies or payloads).
- Images are auto-fit to the screenshot column, and each row is sized to its image.
"""
import json, os, sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
try:
    from PIL import Image as PILImage
    HAVE_PIL = True
except Exception:
    HAVE_PIL = False

DASHES = {"—": "-", "–": "-", "‒": "-", "―": "-", "−": "-"}
ABSENT = "(absent)"


def nodash(s):
    if not isinstance(s, str):
        s = str(s)
    for k, v in DASHES.items():
        s = s.replace(k, v)
    return s


FONT, MONO = "Arial", "Consolas"
HDR = PatternFill("solid", fgColor="1F2A44")
FILLS = {"pass": PatternFill("solid", fgColor="D6EAD6"),
         "fail": PatternFill("solid", fgColor="F7D6D6"),
         "warn": PatternFill("solid", fgColor="FCEFC7"),
         "na":   PatternFill("solid", fgColor="E6E6E6")}
# Keyed on the lowercased `vendor` value. Several spellings map to one colour on
# purpose: the platform question offers "Bing (Microsoft UET)" and "Meta", and a
# vendor that arrives spelled the way the user picked it should still be tinted.
VENDOR_FILLS = {"ga4": "DCE7F5", "meta": "D9E3F7", "facebook": "D9E3F7",
                "meta (facebook)": "D9E3F7",
                "tiktok": "EADCF5", "tik tok": "EADCF5",
                "google ads": "FBE3D6", "floodlight": "FBE3D6",
                "microsoft uet": "DCF0EE", "bing": "DCF0EE", "uet": "DCF0EE",
                "bing (microsoft uet)": "DCF0EE", "microsoft uet (bing)": "DCF0EE",
                "pinterest": "F7D9DE",
                "linkedin": "D9E9F7", "snapchat": "FCF4C7", "segment": "E4EEDC"}
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
IMG_W = 350
LINE_PT = 1.35          # Excel line height as a multiple of font size
IMG_PAD_PT = 14

# Per-platform tab: A event, B conditions, C request, D full payload,
#                   E spec params, F screenshot, G verdict
#
# Columns are WIDE on purpose. The tall content here is a verbatim hit URL and a
# JSON payload, both of which wrap; a narrow column turns one hit into a 40-line
# row and the reader scrolls vertically through a single event. Trading horizontal
# scroll (cheap, one gesture) for row height (expensive, hides the next event) is
# the right way round. Widen these before shortening any evidence cell.
EVENT_COLS = {"A": 20, "B": 34, "C": 76, "D": 64, "E": 44, "F": 54, "G": 52}
EVENT_HEADERS = ["Event name", "Conditions tested", "Request (verbatim)",
                 "Full payload", "Spec parameters", "Location screenshot", "Pass / Fail"]
CONSENT_COLS = {"A": 18, "B": 34, "C": 40, "D": 34, "E": 60, "F": 54}
CONSENT_HEADERS = ["Vendor", "Before consent", "Consent signal", "After consent",
                   "Observation (not a verdict)", "Screenshot"]

# Prose-cell budgets. The audience is analysts who know the vendors and the
# concepts, so these cells are observations, not explanations. Exceeding a budget
# is not fatal (evidence is never truncated), but it is printed at the end of the
# run so an over-written report gets tightened instead of shipped.
MAX_CONDITION = 90          # chars, one line: what was done, not why
MAX_NOTE = 120              # chars per verdict bullet
MAX_NOTES = 5               # bullets per hit
MAX_OBSERVATION = 160       # chars, the Consent tab's observation cell


CHAR_UNITS = 9.5        # see chars_per_line: calibrated, not derived


def chars_per_line(col_width, font_size):
    """How many characters fit on one line of a column of col_width at font_size.

    Excel's width unit is ~1 character of the workbook default font (Calibri 11),
    so a column's capacity in some other font scales by that font's size. The
    9.5 is measured, not derived: an earlier 11.0 overestimated capacity by ~15%,
    every row came out ~2 lines too short, and the tail of each verbatim hit URL
    was clipped where nobody would notice it was missing. If you change a font
    size or a column width, re-export a sheet and check the LAST characters of a
    long URL are still on screen. Everything that needs a column's capacity asks
    this, so the row-height maths and the payload layout cannot drift apart.
    """
    return max(8, int(col_width * CHAR_UNITS / max(font_size, 1)))


def break_pieces(line):
    """Split a logical line where Excel is allowed to wrap it.

    Excel breaks after a space and after a hyphen, and splits mid-token only when
    a token cannot fit a line at all. That matters here because a hit URL is full
    of hyphens (`tid=G-X50M0R6F70`, `en-us`, `tag_exp=115616985~...-118897920`),
    so it wraps ragged: several lines end early at a hyphen, and the URL needs
    more lines than its length divided by the column capacity.
    """
    out, cur = [], ""
    for ch in line:
        cur += ch
        if ch in " -":
            out.append(cur)
            cur = ""
    if cur:
        out.append(cur)
    return out


def wrapped_lines(text, chars):
    """How many display lines `text` occupies in a column `chars` wide.

    Counting "\\n" alone under-counts badly: one verbatim hit URL is a single
    logical line that wraps to seventeen. Counting len/chars under-counts too,
    because of the ragged breaks described in break_pieces, and an under-counted
    row silently clips the TAIL of the URL, which is exactly where the event
    params live. So pack the break pieces the way Excel does.
    """
    total = 0
    for logical in str(text).split("\n"):
        lines, used = 1, 0
        for p in break_pieces(logical):
            if len(p) > chars:                      # too long to fit any line
                if used:
                    lines += 1
                lines += -(-len(p) // chars) - 1    # ceil, minus the line we are on
                used = len(p) % chars or chars
            elif used + len(p) <= chars:
                used += len(p)
            else:
                lines += 1
                used = len(p)
        total += lines
    return total


def cell_height(text, col_width, font_size):
    """Points needed to show `text` wrapped in a column of col_width at font_size.

    Line height comes off the font size too, so the two cannot drift apart.
    """
    if not text:
        return LINE_PT * font_size
    return wrapped_lines(text, chars_per_line(col_width, font_size)) * font_size * LINE_PT


def row_height(cells, img_px=0, floor=0):
    """Tallest of the wrapped cells and the image, with a floor. cells = (text, width, font)."""
    need = [cell_height(t, w, f) for t, w, f in cells]
    need.append(img_px * 0.75 + IMG_PAD_PT)     # px -> pt, so the row fits the image
    need.append(floor)
    return max(need)


def header(ws, headers, widths, height=28):
    for k, v in widths.items():
        ws.column_dimensions[k].width = v
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
        cell.fill = HDR
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = height
    ws.freeze_panes = "A2"


def place_image(ws, row, col_letter, shots_dir, img_file):
    """Add an image scaled to the column. Returns its height in px (0 if none)."""
    if not img_file or str(img_file).strip().lower() in ("none", "null", ""):
        return 0
    path = os.path.join(shots_dir, img_file)
    if not (os.path.exists(path) and HAVE_PIL):
        return 0
    with PILImage.open(path) as im:
        w0, h0 = im.size
    img = XLImage(path)
    img.width = IMG_W
    img.height = int(IMG_W * h0 / w0)
    ws.add_image(img, "%s%d" % (col_letter, row))
    return img.height


# ---------------------------------------------------------------- payload cell

SCALAR = (str, int, float, bool, type(None))


def payload_text(payload, col_width, font_size):
    """The Full payload cell: every param, values verbatim, laid out to be short.

    One param per line is what actually makes these rows taller than the screen:
    a GA4 hit carries 40+ params, so a JSON dump is a 45-line cell no column
    width can shrink. Scalar params are therefore packed two per line when both
    halves fit, which roughly halves the tallest cell in most rows. The payload
    font is monospace, so the two columns line up.

    Nothing is altered: no value is truncated, reordered or reformatted, only
    padded with spaces. Values go through json.dumps so the string "2" still
    reads differently from the number 2, which is itself a finding in this QA
    (`ep.value` vs `epn.value`). Nested values (GA4 items, a JSON body) keep a
    full-width JSON block, because their structure is the readable part.
    """
    if not isinstance(payload, dict):
        return json.dumps(payload, indent=2, ensure_ascii=False)

    flat, nested = [], []
    for k, v in payload.items():
        if isinstance(v, SCALAR):
            flat.append("%s: %s" % (k, json.dumps(v, ensure_ascii=False)))
        else:
            nested.append('%s: %s' % (k, json.dumps(v, indent=2, ensure_ascii=False)))

    half = (chars_per_line(col_width, font_size) - 2) // 2
    lines, i = [], 0
    while i < len(flat):
        a = flat[i]
        b = flat[i + 1] if i + 1 < len(flat) else None
        if b is not None and len(a) <= half and len(b) <= half:
            lines.append(a.ljust(half + 2) + b)
            i += 2
        else:
            lines.append(a)
            i += 1
    # An empty dict is not the same as no payload: say so rather than render blank.
    return "\n".join(lines + nested) or "(payload decoded to no params)"


# ---------------------------------------------------------------- spec params

def _item_lookup(item, key):
    """Item dicts are keyed 'item_name (nm)', so match on the readable half too."""
    if not isinstance(item, dict):
        return ABSENT
    if key in item:
        return item[key]
    kl = key.strip().lower()
    for k, v in item.items():
        if k.strip().lower() == kl or k.split(" (")[0].strip().lower() == kl:
            return v
    return ABSENT


def lookup(payload, path):
    """Resolve one spec param against the decoded payload.

    Literal key FIRST, because vendor params legitimately contain dots
    ("epn.value", "ep.transaction_id") and splitting them would lose the value.
    """
    if not isinstance(payload, dict):
        return "(no payload captured)"
    if path in payload:
        return payload[path]
    if "items[]." in path:
        key = path.split("items[].", 1)[1]
        items = payload.get("items")
        if not isinstance(items, list) or not items:
            return ABSENT
        vals = [_item_lookup(it, key) for it in items]
        return vals[0] if len(vals) == 1 else vals
    cur = payload
    for part in path.split("."):
        if isinstance(cur, dict) and part in cur:
            cur = cur[part]
        else:
            return ABSENT
    return cur


def spec_view(ev):
    """The 'Spec parameters' cell: only what the spec asked about.

    Derived from ev["payload"] so it can never contradict the Full payload column.
    """
    if ev.get("spec_payload") is not None:
        return ev["spec_payload"]
    params = ev.get("spec_params")
    if not params:
        return None
    return {p: lookup(ev.get("payload"), p) for p in params}


# ---------------------------------------------------------------- brevity

def check_prose(label, wordy, **cells):
    """Note any prose cell that overruns its budget. Never rewrites the text.

    Only the cells the model writes are measured. The verbatim URL, body, payload
    and absence evidence have no budget: they are the evidence, and shortening
    them to fit a column is the one thing this report must never do.
    """
    for kind, value in cells.items():
        if kind == "notes":
            notes = value if isinstance(value, list) else ([value] if value else [])
            if len(notes) > MAX_NOTES:
                wordy.append("%s  %d verdict bullets (budget %d)"
                             % (label, len(notes), MAX_NOTES))
            for n in notes:
                if len(str(n)) > MAX_NOTE:
                    wordy.append("%s  verdict bullet %d chars (budget %d): %s..."
                                 % (label, len(str(n)), MAX_NOTE, str(n)[:60]))
            continue
        budget = {"conditions": MAX_CONDITION, "observation": MAX_OBSERVATION}[kind]
        if value and len(str(value)) > budget:
            wordy.append("%s  %s %d chars (budget %d)"
                         % (label, kind, len(str(value)), budget))


# ---------------------------------------------------------------- sheets

def sheet_name(vendor, used):
    """Excel: <=31 chars, none of : \\ / ? * [ ], non-blank, unique."""
    s = "".join(c for c in str(vendor) if c not in set(':\\/?*[]')).strip() or "Unnamed"
    s = s[:31]
    base, n = s, 2
    while s.lower() in used:
        suffix = " (%d)" % n
        s = base[:31 - len(suffix)] + suffix
        n += 1
    used.add(s.lower())
    return s


def request_cell(ev):
    """method + endpoint + the verbatim URL and body. Verbatim, never trimmed.

    An absence claim must carry its evidence. A run reported "no pre-consent hits"
    and "view_item_list never fired" when both were plainly in the network log,
    because a filtered read came back empty and the null was written up as a
    finding. Prose telling the model to check harder did not prevent it, so this
    is a gate (ADR-0006): no `absence_evidence`, no clean-looking cell.
    """
    if not ev.get("sent", True):
        ev_txt = str(ev.get("absence_evidence", "")).strip()
        if not ev_txt:
            return ("No request observed\n\n"
                    "!! UNVERIFIED: no absence evidence was recorded for this row.\n"
                    "Re-check with window.__nqaRequests() and paste the total number of\n"
                    "requests searched, or delete this row. Do not hand this over as is.")
        return "No request observed\n\nEvidence:\n" + ev_txt
    parts = ["%s %s" % (ev.get("method", "GET"), ev.get("endpoint", ""))]
    if ev.get("count") and int(ev["count"]) > 1:
        parts.append("fired %d times for this one interaction" % int(ev["count"]))
    if ev.get("url"):
        parts += ["", "URL:", ev["url"]]
    if ev.get("body"):
        parts += ["", "Body:", ev["body"]]
    return "\n".join(parts)


def vendor_sheet(wb, vendor, events, shots_dir, used_names, unverified, wordy):
    ws = wb.create_sheet(sheet_name(vendor, used_names))
    header(ws, EVENT_HEADERS, EVENT_COLS)
    tab = VENDOR_FILLS.get(str(vendor).strip().lower())
    if tab:
        ws.sheet_properties.tabColor = tab

    r = 2
    for ev in events:
        payload = ev.get("payload", None)
        if payload is not None:
            pj = payload_text(payload, EVENT_COLS["D"], 8)           # values verbatim
        else:
            pj = nodash(ev.get("payload_note", "(no payload captured)"))

        sv = spec_view(ev)
        sj = ("(no spec parameters given for this event)" if sv is None
              else json.dumps(sv, indent=2, ensure_ascii=False))

        req = request_cell(ev)
        notes = ev.get("notes", [])
        notes_txt = nodash("\n\n".join(notes) if isinstance(notes, list) else str(notes))
        cond = nodash(ev.get("conditions", ""))
        check_prose("%s / %s" % (vendor, ev.get("event", "")), wordy,
                    conditions=cond, notes=notes)

        ws.cell(row=r, column=1, value=nodash(ev.get("event", ""))).font = Font(name=FONT, size=11, bold=True)
        ws.cell(row=r, column=2, value=cond).font = Font(name=FONT, size=10)
        ws.cell(row=r, column=3, value=req).font = Font(name=MONO, size=8)
        ws.cell(row=r, column=4, value=pj).font = Font(name=MONO, size=8)
        ws.cell(row=r, column=5, value=sj).font = Font(name=MONO, size=9)
        ws.cell(row=r, column=7, value=notes_txt).font = Font(name=FONT, size=10)
        for c in range(1, 8):
            cc = ws.cell(row=r, column=c)
            cc.alignment = TOP
            cc.border = BORDER

        vfill = FILLS.get(ev.get("verdict", "warn"))
        ws.cell(row=r, column=1).fill = vfill
        ws.cell(row=r, column=7).fill = vfill
        # A spec param the hit is missing is the point of the column: tint it.
        if sv and ABSENT in [str(v) for v in sv.values()]:
            ws.cell(row=r, column=5).fill = FILLS["fail"]
        if req.startswith("No request observed") and "!! UNVERIFIED" in req:
            ws.cell(row=r, column=3).fill = FILLS["warn"]
            unverified.append("%s / %s" % (vendor, ev.get("event", "")))

        img_h = place_image(ws, r, "F", shots_dir, ev.get("location_image"))
        ws.row_dimensions[r].height = row_height(
            [(req, EVENT_COLS["C"], 8), (pj, EVENT_COLS["D"], 8),
             (sj, EVENT_COLS["E"], 9), (notes_txt, EVENT_COLS["G"], 10),
             (cond, EVENT_COLS["B"], 10)],
            img_px=img_h, floor=120)
        r += 1
    return r - 2


def consent_sheet(ws, consent, shots_dir, unverified_consent, wordy):
    """First tab. Descriptive only. There is deliberately NO pass/fail column."""
    ws.title = "Consent"
    header(ws, CONSENT_HEADERS, CONSENT_COLS)

    rows = [c for c in consent if "vendor" in c]
    meta = next((c for c in consent if c.get("state")), {})
    state = meta.get("state")

    r = 2
    if state:
        if state == "forced":
            note = ("Pre-consent state was forced: site cookies and web storage were "
                    "cleared and the page reloaded, so the rows below compare each "
                    "vendor before and after consent was granted.")
        elif state == "already_accepted":
            note = ("Consent was already accepted and could not be reset, so "
                    "pre-consent behaviour was not observed.")
        else:
            note = ("Consent had not been given when the session started, so the rows "
                    "below compare behaviour before and after granting it.")
        if meta.get("reset"):
            note += " " + str(meta["reset"])
        cell = ws.cell(row=r, column=1, value=nodash(note))
        cell.font = Font(name=FONT, size=10, italic=True)
        cell.alignment = TOP
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.row_dimensions[r].height = row_height(
            [(nodash(note), sum(CONSENT_COLS.values()), 10)], floor=30)
        r += 1

    for c in rows:
        before = str(c.get("before", ""))
        # "no hits before consent" is the exact claim a run got wrong, so it has
        # to carry evidence too (ADR-0006 gate, not more prose).
        if (any(w in before.lower() for w in ("no hit", "none", "no request", "nothing"))
                and not str(c.get("evidence", "")).strip()):
            before += ("\n\n!! UNVERIFIED: re-check with window.__nqaRequests() and "
                       "record how many requests were searched.")
            unverified_consent.append(str(c.get("vendor", "")))
        elif str(c.get("evidence", "")).strip():
            before += "\n\n" + str(c["evidence"])
        vals = [c.get("vendor", ""), before, c.get("signal", ""),
                c.get("after", ""), c.get("observation", "")]
        check_prose("Consent / %s" % c.get("vendor", ""), wordy,
                    observation=c.get("observation", ""))
        for i, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=i, value=nodash(v))
            cell.font = Font(name=FONT, size=10, bold=(i == 1))
        for i in range(1, 7):
            cc = ws.cell(row=r, column=i)
            cc.alignment = TOP
            cc.border = BORDER
        vf = VENDOR_FILLS.get(str(c.get("vendor", "")).strip().lower())
        if vf:
            ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=vf)
        if "!! UNVERIFIED" in before:
            ws.cell(row=r, column=2).fill = FILLS["warn"]

        img_h = place_image(ws, r, "F", shots_dir, c.get("location_image"))
        widths = [CONSENT_COLS[k] for k in ("A", "B", "C", "D", "E")]
        ws.row_dimensions[r].height = row_height(
            [(nodash(v), w, 10) for v, w in zip(vals, widths)],
            img_px=img_h, floor=60)
        r += 1
    return len(rows)


def main():
    if len(sys.argv) < 5:
        print(__doc__)
        sys.exit(1)
    events_path, consent_path, shots_dir, out_path = sys.argv[1:5]
    with open(events_path, encoding="utf-8") as f:
        events = json.load(f)
    consent = []
    if consent_path.lower() not in ("none", "-", ""):
        with open(consent_path, encoding="utf-8") as f:
            consent = json.load(f)

    # group by vendor, keeping first-appearance order so the funnel story survives
    order, by_vendor = [], {}
    for ev in events:
        v = ev.get("vendor") or "Unclassified"
        if v not in by_vendor:
            order.append(v)
            by_vendor[v] = []
        by_vendor[v].append(ev)

    wb = openpyxl.Workbook()
    used = set()
    unverified_consent = []
    wordy = []
    n_consent = consent_sheet(wb.active, consent, shots_dir, unverified_consent, wordy)  # tab 1
    used.add("consent")
    unverified = []
    counts = [(v, vendor_sheet(wb, v, by_vendor[v], shots_dir, used, unverified, wordy))
              for v in order]
    wb.save(out_path)

    if len(os.path.abspath(out_path)) >= 259:
        print("WARNING: output path is >= 259 chars, Windows Excel may refuse to open it. "
              "Use a shorter filename.")
    print("saved %s" % out_path)
    print("  Consent rows %d" % n_consent)
    for v, n in counts:
        print("  %-16s %d hits" % (v, n))
    if unverified or unverified_consent:
        print("")
        print("  !! %d absence claim(s) carry NO evidence and are marked UNVERIFIED "
              "in the workbook:" % (len(unverified) + len(unverified_consent)))
        for u in unverified:
            print("       hit row     %s" % u)
        for u in unverified_consent:
            print("       consent row %s" % u)
        print("     Re-check each with window.__nqaRequests() before handing this over.")
    if wordy:
        print("")
        print("  %d prose cell(s) are longer than the report's style allows. The reader is an"
              % len(wordy))
        print("  analyst who knows these vendors: cut to observation only, no explanation.")
        for w in wordy:
            print("       %s" % w)


if __name__ == "__main__":
    main()
